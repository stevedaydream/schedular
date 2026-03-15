#!/usr/bin/env python3
"""
parse_schedule_excel.py
解析 9A 值班 Excel → historical_import.json

班別對應：
  D / d         → D
  N / n         → N
  OFF / off     → Off
  S1 / 9A / 9a → S1  (原 AM 班)
  9B            → Off (支援他科 = 本科不排班)
  H3            → H3  (限定星期六上班日)
  S1D / NrsD    → D   (支援+D 以 D 計)
  Nr            → S1
  其他           → null (不匯入)

整排空白 (col 1 = 'X') = 留職停薪/育嬰/離職，不匯入任何班別。
"""

import openpyxl
import json
import re
from datetime import datetime

# ── 設定 ──────────────────────────────────────────────────────
EXCEL_PATH = r"G:\project\scheduler\tools\schedule_temp.xlsm"
OUTPUT_PATH = r"G:\project\scheduler\tools\historical_import.json"

SHIFT_MAP = {
    'D': 'D', 'd': 'D',
    'N': 'N', 'n': 'N',
    'OFF': 'Off', 'Off': 'Off', 'off': 'Off',
    'S1':  'S1', '9A': 'S1', '9a': 'S1',
    '9B':  'Off',
    'H3':  'H3',
    'S1D': 'D', 'NrsD': 'D',
    'Nr':  'S1',
    'D/N': 'D', 'Cover': 'D',
    'W6': 'Off', 'W6 OFF': 'Off',
}

# 整排空白的標記值
WHOLE_MONTH_BLANK_MARKERS = {'X', 'x'}

# 有效代號範圍（單個大寫字母）
VALID_CODES = set('ABCDEFGHIJKLMNOPQRSTUVWXYZ')

# 這些字串開頭是大寫字母但不是人名
EXCLUDE_NAMES = {'D', 'N', 'OFF', 'NURSE', 'Cover', 'Ortho', 'Plasty', 'Uro', 'ACLS', 'S1', 'S', 'H3', 'W6'}


def map_shift(val):
    if val is None:
        return None
    v = str(val).strip()
    if v == '' or v == '0':
        return None
    if v in SHIFT_MAP:
        return SHIFT_MAP[v]
    if v.upper() in SHIFT_MAP:
        return SHIFT_MAP[v.upper()]
    return None


def sheet_to_yyyymm(sheet_name):
    """Convert sheet name to yyyyMM string. Returns None if unrecognisable."""
    name = sheet_name.strip()
    # Remove suffixes: " (2)", "-1", etc.
    clean = re.sub(r'\s*\([^)]*\)', '', name)
    clean = re.sub(r'[-]\d+$', '', clean)
    clean = clean.strip()

    if not clean.isdigit():
        return None

    if len(clean) == 5:
        first3 = int(clean[:3])
        if 100 <= first3 < 200:
            # ROC year: e.g. 11306 → 113 + 1911 = 2024, month 06
            greg_year = first3 + 1911
            month = int(clean[3:])
        else:
            # Short Gregorian: e.g. 20251 → 2025, 1
            greg_year = int(clean[:4])
            month = int(clean[4:])
        if 1 <= month <= 12:
            return f"{greg_year}{month:02d}"
        return None

    if len(clean) == 6:
        # e.g. 202601
        year = int(clean[:4])
        month = int(clean[4:])
        if 2000 <= year <= 2100 and 1 <= month <= 12:
            return clean
        return None

    return None


def find_date_row(rows):
    """Return (row_index, {col_index: day}) for the row with ≥20 date cells."""
    for i, row in enumerate(rows):
        day_cols = {}
        for ci, v in enumerate(row):
            if isinstance(v, datetime):
                day_cols[ci] = v.day
        if len(day_cols) >= 20:
            return i, day_cols
    return None, {}


def is_person_name(val):
    """Return True if val looks like a person name (starts with single uppercase letter)."""
    if not isinstance(val, str) or len(val) < 1:
        return False
    first = val[0]
    if first not in VALID_CODES:
        return False
    # Must not be a known keyword
    if val.strip() in EXCLUDE_NAMES:
        return False
    # Should be more than 1 character (code + name), unless it's literally just a single letter row
    return True


def parse_sheet(ws, sheet_name):
    """Parse one worksheet. Returns dict: { code: { day: shift } } or None on failure."""
    rows = list(ws.iter_rows(values_only=True))
    date_row_idx, day_cols = find_date_row(rows)
    if date_row_idx is None or not day_cols:
        return None

    col_start = min(day_cols.keys())
    col_end = max(day_cols.keys())

    people = {}

    # Person rows start 2 rows after the date row
    for i in range(date_row_idx + 2, min(date_row_idx + 30, len(rows))):
        row = rows[i]
        if not row:
            continue

        # Try col 0 and col 1 for person name
        name = None
        name_col = None
        for nc in [0, 1]:
            v = row[nc] if len(row) > nc else None
            if is_person_name(v):
                # Exclude names that look like shift values or pure stats
                if not any(row[c] for c in day_cols if c < len(row) and isinstance(row[c], (int, float)) and not isinstance(row[c], bool)):
                    name = v
                    name_col = nc
                    break
                else:
                    name = v
                    name_col = nc
                    break

        if name is None:
            continue

        code = name[0].upper()

        # Check for whole-month blank marker: col 1 (if name in col 0) or col 2 (if name in col 1)
        check_col = 1 if name_col == 0 else 2
        blank_marker = row[check_col] if len(row) > check_col else None
        if blank_marker in WHOLE_MONTH_BLANK_MARKERS:
            # Whole month absent — record but no shifts
            if code not in people:
                people[code] = {}
            continue

        # Extract shifts for each day
        shifts = {}
        valid_shift_count = 0
        for ci, day in day_cols.items():
            val = row[ci] if ci < len(row) else None
            mapped = map_shift(val)
            if mapped is not None:
                shifts[day] = mapped
                valid_shift_count += 1

        # Skip rows that have no valid shifts AND no marker (likely totals/summary rows)
        if valid_shift_count == 0 and code not in people:
            continue

        if code not in people:
            people[code] = {}
        people[code].update(shifts)

    return people if people else None


def parse_excel(filepath):
    print(f"開啟：{filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=False, data_only=True)

    result = {}
    skipped = []

    for sheet_name in wb.sheetnames:
        yyyyMM = sheet_to_yyyymm(sheet_name)
        if not yyyyMM:
            skipped.append(f"{sheet_name} (無法轉換月份)")
            continue

        ws = wb[sheet_name]
        people = parse_sheet(ws, sheet_name)

        if people is None:
            skipped.append(f"{sheet_name} (找不到日期列)")
            continue

        # Convert to output format: { code: { "day_N": shift } }
        month_data = {}
        for code, shifts in people.items():
            day_shifts = {f"day_{day}": shift for day, shift in shifts.items()}
            month_data[code] = day_shifts

        if month_data:
            result[yyyyMM] = month_data
            shift_count = sum(len(s) for s in month_data.values())
            print(f"  {sheet_name:15s} → {yyyyMM}: {len(month_data)} 人, {shift_count} 格班別")
        else:
            skipped.append(f"{sheet_name} (無有效人員資料)")

    if skipped:
        print(f"\n略過 {len(skipped)} 個分頁：")
        for s in skipped:
            print(f"  - {s}")

    return result


def main():
    print("=" * 60)
    print("9A 值班表歷史資料解析工具")
    print("=" * 60)

    data = parse_excel(EXCEL_PATH)

    # Collect all codes across all months
    all_codes = set()
    for month_data in data.values():
        all_codes.update(month_data.keys())

    output = {
        "source": EXCEL_PATH,
        "generatedAt": datetime.now().isoformat(),
        "codes": sorted(all_codes),
        "months": data
    }

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"解析完成！")
    print(f"  月份：{len(data)} 個月")
    print(f"  代號：{sorted(all_codes)}")
    print(f"  輸出：{OUTPUT_PATH}")
    print(f"{'='*60}")
    print("\n下一步：在系統管理 → 備份匯出頁面，選擇「匯入歷史班表」上傳此 JSON 檔案。")


if __name__ == '__main__':
    main()
